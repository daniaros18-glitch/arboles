# -*- coding: utf-8 -*-
"""Genera la carta Gantt y el presupuesto del Fondo FMA sobre las planillas oficiales.

Uso:  python scripts/generar_planillas_fma.py
Lee   __act.json (actividades extraidas de FMA_2026_ACTIVIDADES.md)
Sale  postulaciones/entrega/FMA_Carta_Gantt.xlsx y FMA_Presupuesto.xlsx
"""
import io
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")
SALIDA = os.path.join(BASE, "postulaciones", "entrega")
os.makedirs(SALIDA, exist_ok=True)

TITULO = ("¿Qué árboles sostienen más vida? Línea base ciudadana del arbolado "
          "urbano de Temuco")
DIRECCION = "[PENDIENTE: dirección registrada de Acción Ecologista Ekuwün]"
FECHA = "23 de agosto de 2026"

VERDE = PatternFill("solid", fgColor="1B4D3E")
VERDE_CLARO = PatternFill("solid", fgColor="D6E4DE")
GRIS = PatternFill("solid", fgColor="EFEFEF")
BORDE = Border(*[Side(style="thin", color="BBBBBB")] * 4)


def desagrupar(ws):
    """La plantilla oficial trae celdas combinadas que impiden escribir."""
    for rango in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rango))


# ----------------------------------------------------------------- CARTA GANTT
def carta_gantt():
    acts = json.load(io.open(os.path.join(BASE, "scripts", "actividades_gantt.json"), encoding="utf-8"))
    wb = openpyxl.load_workbook(os.path.join(DOWNLOADS, "Planilla_tipo_carta_gantt.xlsx"))
    ws = wb.active
    desagrupar(ws)

    ws["A4"] = "Nombre del proyecto:  " + TITULO

    # 9 meses de ejecucion, 36 semanas. La plantilla trae 13 meses de junio a junio.
    meses = ["DICIEMBRE 2026", "ENERO 2027", "FEBRERO 2027", "MARZO 2027", "ABRIL 2027",
             "MAYO 2027", "JUNIO 2027", "JULIO 2027", "AGOSTO 2027"]
    for i, m in enumerate(meses):
        ws.cell(row=5, column=2 + i * 4).value = m
    # limpiar los 4 meses sobrantes de la plantilla y sus semanas
    for col in range(2 + 9 * 4, 64):
        ws.cell(row=5, column=col).value = None
        ws.cell(row=6, column=col).value = None

    etapas = {
        "1": "ETAPA 1  ·  Preparación y formación   (dic 2026 a feb 2027)",
        "2": "ETAPA 2  ·  Terreno con las comunidades escolares   (mar a abr 2027)",
        "3": "ETAPA 3  ·  Validación, análisis y componente artístico   (may a jul 2027)",
        "4": "ETAPA 4  ·  Devolución, sistema de monitoreo e incidencia   (jul a ago 2027)",
    }

    # limpiar el area de actividades de la plantilla
    for r in range(7, 60):
        for c in range(1, 64):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill()

    fila = 7
    etapa_actual = None
    for a in acts:
        et = a["id"].split(".")[0]
        if et != etapa_actual:
            etapa_actual = et
            cel = ws.cell(row=fila, column=1)
            cel.value = etapas[et]
            cel.font = Font(bold=True, color="FFFFFF", size=10)
            cel.fill = VERDE
            for c in range(2, 38):
                ws.cell(row=fila, column=c).fill = VERDE
            fila += 1

        ws.cell(row=fila, column=1).value = "  " + a["id"] + "  " + a["d"]
        ws.cell(row=fila, column=1).font = Font(size=9)
        ws.cell(row=fila, column=1).alignment = Alignment(vertical="center", wrap_text=False)
        for s in range(1, 37):
            cel = ws.cell(row=fila, column=1 + s)
            cel.border = BORDE
            if a["a"] <= s <= a["b"]:
                cel.fill = VERDE_CLARO
        fila += 1

    ws.column_dimensions["A"].width = 78
    for s in range(1, 37):
        ws.column_dimensions[get_column_letter(1 + s)].width = 3.1

    nota = fila + 1
    ws.cell(row=nota, column=1).value = (
        "Los meses de la plantilla original (junio a junio) fueron reetiquetados a la ventana "
        "de ejecución del fondo: diciembre 2026 a agosto 2027, 9 meses y 36 semanas.")
    ws.cell(row=nota, column=1).font = Font(size=8, italic=True, color="666666")

    ruta = os.path.join(SALIDA, "FMA_Carta_Gantt.xlsx")
    wb.save(ruta)
    return ruta, len(acts), fila - 7


# ----------------------------------------------------------------- PRESUPUESTO
# Honorarios: valor LIQUIDO mensual o por jornada. La retencion se suma al costo.
HONORARIOS = [
    ("Coordinación general", "mes", 9, 110000),
    ("Análisis territorial, datos y análisis estadístico", "mes", 5, 110000),
    ("Especialista en ciencias naturales (validación taxonómica)", "mes", 4, 140000),
    ("Coordinación pedagógica", "mes", 4, 140000),
    ("Terreno y control de calidad (submuestra ciega)", "jornada", 8, 45000),
    ("Diseño gráfico del sistema de fichas", "producto", 1, 130000),
]
PRODUCCION = [
    ("Kits de ciencia ciudadana: huinchas, varas de 1,30 m, clinómetros, grillas, "
     "clips macro, lupas, cronómetros, portapapeles", "kit", 4, 88000),
    ("Óptica: 6 monoculares y 1 binocular", "global", 1, 195000),
    ("Seguridad: 60 chalecos reflectantes infantiles y 4 botiquines", "global", 1, 260000),
    ("Traslados del equipo a las tres escuelas", "mes", 9, 15000),
    ("Transporte escolar, salida al Humedal Urbano Antumalén", "salida", 1, 120000),
    ("Alimentación de la jornada de formación docente", "jornada", 1, 80000),
    ("Convocatoria y materiales de difusión de las devoluciones barriales", "global", 1, 50000),
    ("Colación de las tres devoluciones barriales y del encuentro de cierre", "instancia", 4, 35000),
    ("Arriendo de amplificación y proyección para las instancias de devolución", "global", 1, 50000),
    ("Transporte de las delegaciones escolares al encuentro de cierre", "global", 1, 70000),
]
# OJO: las bases excluyen "edicion e impresion de publicaciones". Estas partidas son
# material de trabajo de terreno e instalacion de capacidades, no productos editoriales,
# y estan redactadas para que se lea asi.
OTROS = [
    ("Producción del material de terreno: fichas F1 en tres formatos, F2, F3 y "
     "protocolo, plastificados para uso en terreno", "global", 1, 180000),
    ("Materiales del herbario ilustrado elaborado por las y los estudiantes", "global", 1, 85000),
    ("Producción del cuadernillo docente y del kit de ciencia ciudadana, para "
     "instalación de capacidades en las tres escuelas", "global", 1, 150000),
    ("Producción de los expedientes de árbol patrimonial (Art. 9) y de las fichas "
     "de árbol que se entregan a cada comunidad", "global", 1, 40000),
    ("Producción de los resultados para cada barrio: panel de datos del sector y "
     "fichas de los árboles destacados, que quedan en la escuela y en la junta de vecinos",
     "global", 1, 100000),
]
RET = 0.145
IVA = 0.19


def presupuesto():
    wb = openpyxl.load_workbook(os.path.join(DOWNLOADS, "Planilla_tipo_de_presupuesto.xlsx"))
    ws = wb.active
    desagrupar(ws)
    ws["D6"] = TITULO
    ws["D7"] = DIRECCION
    ws["D8"] = FECHA

    for r in range(15, 60):
        for c in range(3, 13):
            ws.cell(row=r, column=c).value = None

    fila = 15

    def bloque(nombre, filas, con_iva):
        nonlocal fila
        ini = fila
        for i, (partida, unidad, cant, unit) in enumerate(filas):
            ws.cell(row=fila, column=3).value = nombre if i == 0 else None
            ws.cell(row=fila, column=4).value = partida
            ws.cell(row=fila, column=5).value = 1
            ws.cell(row=fila, column=6).value = cant
            ws.cell(row=fila, column=7).value = unit
            ws.cell(row=fila, column=8).value = "=G{0}*E{0}*F{0}".format(fila)
            ws.cell(row=fila, column=9).value = "=H{0}*{1}".format(fila, IVA) if con_iva else 0
            ws.cell(row=fila, column=10).value = 0 if con_iva else "=H{0}*{1}".format(fila, RET)
            ws.cell(row=fila, column=11).value = "=H{0}+I{0}+J{0}".format(fila)
            ws.cell(row=fila, column=12).value = "Fondo FMA"
            ws.cell(row=fila, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            for c in range(3, 13):
                ws.cell(row=fila, column=c).border = BORDE
                if c >= 7:
                    ws.cell(row=fila, column=c).number_format = "#,##0"
            fila += 1
        # subtotal del item
        ws.cell(row=fila, column=4).value = "Subtotal " + nombre.split("(")[0].strip()
        ws.cell(row=fila, column=4).font = Font(bold=True)
        for c, letra in ((8, "H"), (9, "I"), (10, "J"), (11, "K")):
            ws.cell(row=fila, column=c).value = "=SUM({0}{1}:{0}{2})".format(letra, ini, fila - 1)
            ws.cell(row=fila, column=c).font = Font(bold=True)
            ws.cell(row=fila, column=c).number_format = "#,##0"
        for c in range(3, 13):
            ws.cell(row=fila, column=c).fill = GRIS
            ws.cell(row=fila, column=c).border = BORDE
        sub = fila
        fila += 2
        return sub

    s1 = bloque("1. Honorarios de equipo o asesorías", HONORARIOS, con_iva=False)
    s2 = bloque("2. Producción (traslados, montaje, arriendos y compras de equipos)",
                PRODUCCION, con_iva=True)
    s3 = bloque("3. Otros", OTROS, con_iva=True)

    ws.cell(row=fila, column=4).value = "TOTAL DE PROYECTO"
    ws.cell(row=fila, column=11).value = "=K{0}+K{1}+K{2}".format(s1, s2, s3)
    ws.cell(row=fila + 1, column=4).value = "TOTAL A POSTULAR FONDO FMA"
    ws.cell(row=fila + 1, column=11).value = "=K{0}".format(fila)
    for r in (fila, fila + 1):
        ws.cell(row=r, column=4).font = Font(bold=True, size=11)
        ws.cell(row=r, column=11).font = Font(bold=True, size=11, color="1B4D3E")
        ws.cell(row=r, column=11).number_format = "#,##0"
        for c in range(3, 13):
            ws.cell(row=r, column=c).border = BORDE

    n = fila + 4
    ws.cell(row=n, column=4).value = "_______________________________________"
    ws.cell(row=n + 1, column=4).value = "Firma del postulante o representante legal"
    ws.cell(row=n + 2, column=4).value = "[PENDIENTE: nombre del representante legal de Acción Ecologista Ekuwün]"
    ws.cell(row=n + 4, column=4).value = (
        "Notas: montos en pesos chilenos. Los honorarios se expresan como valor líquido "
        "y la retención de 14,5% se suma al costo del proyecto, siguiendo el encabezado "
        "de la planilla oficial. El IVA de 19% aplica a las compras de los ítems 2 y 3.")
    ws.cell(row=n + 4, column=4).font = Font(size=8, italic=True, color="666666")
    ws.cell(row=n + 4, column=4).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 52
    for col in "EFGHIJK":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["L"].width = 18

    ruta = os.path.join(SALIDA, "FMA_Presupuesto.xlsx")
    wb.save(ruta)

    tot = 0
    for filas, con_iva in ((HONORARIOS, False), (PRODUCCION, True), (OTROS, True)):
        for _, _, cant, unit in filas:
            h = cant * unit
            tot += h * (1 + (IVA if con_iva else RET))
    return ruta, round(tot)


if __name__ == "__main__":
    g, nact, nfilas = carta_gantt()
    p, total = presupuesto()
    print("Gantt:", g, "|", nact, "actividades en", nfilas, "filas")
    print("Presupuesto:", p, "| TOTAL =", format(total, ",d").replace(",", "."))
    print("Holgura respecto de 6.000.000:", format(6000000 - total, ",d").replace(",", "."))
